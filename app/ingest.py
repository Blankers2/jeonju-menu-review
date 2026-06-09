"""엑셀 입력 파싱: 이미지 마스터(place_item_list)와 이미지별 번역본."""
import io
import re
from pathlib import Path

import openpyxl


def _norm(s) -> str:
    return str(s).strip() if s is not None else ""


# 파일명 예: place_13764_image_111225_20260512_111947911.xlsx
_FNAME_RE = re.compile(r"place[_-](\d+)[_-]image[_-](\d+)", re.IGNORECASE)
# 폴더명 예: 13764_전주성갈비  (place_id + 가게명)
_FOLDER_RE = re.compile(r"^(\d+)[ _\-]+(.+)$")


def parse_path_meta(rel_path: str) -> dict:
    """업로드 파일의 상대경로 -> {place_id, item_id, title}.

    place_id/item_id는 파일명에서, title(가게명)은 상위폴더명(`<place_id>_<가게명>`)에서.
    """
    parts = str(rel_path).replace("\\", "/").split("/")
    fname = parts[-1] if parts else ""
    place_id = item_id = None
    title = ""
    m = _FNAME_RE.search(fname)
    if m:
        place_id = int(m.group(1))
        item_id = m.group(2)
    if len(parts) >= 2:
        fm = _FOLDER_RE.match(parts[-2].strip())
        if fm:
            if place_id is None:
                place_id = int(fm.group(1))
            title = fm.group(2).strip()
    return {"place_id": place_id, "item_id": item_id, "title": title}


def load_images(path: Path) -> dict:
    """place_item_list.xlsx -> {item_id(str): {place_id, title, image_url, width, height}}.

    헤더: place_id, item_id, title, GA Code, image_width, image_height, item_status, image_url
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [_norm(h) for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    out = {}
    for row in rows:
        item_id = cell(row, "item_id")
        if item_id is None or _norm(item_id) == "":
            continue
        item_id = str(int(item_id)) if isinstance(item_id, float) else str(item_id).strip()
        place_id = cell(row, "place_id")
        place_id = int(place_id) if place_id is not None and _norm(place_id) != "" else None
        out[item_id] = {
            "place_id": place_id,
            "title": _norm(cell(row, "title")),
            "image_url": _norm(cell(row, "image_url")),
            "width": cell(row, "image_width"),
            "height": cell(row, "image_height"),
        }
    wb.close()
    return out


# 번역본 헤더 매핑 (접두사/키워드로 견고하게 식별)
def _resolve_translation_columns(header: list[str]) -> dict:
    """헤더 리스트 -> {field: col_index}. field in item_id, org_id, ko, en, ja, zh_cn, zh_tw."""
    col = {}
    for i, h in enumerate(header):
        hl = h.lower()
        if h == "Item Id":
            col["item_id"] = i
        elif h == "Item Org Id":
            col["org_id"] = i
        elif h == "Org Content":
            col["ko"] = i
        elif h.startswith("11_") or "simplified" in hl or "简体" in h:
            col["zh_cn"] = i
        elif h.startswith("12_") or "traditional" in hl or "繁體" in h or "繁体" in h:
            col["zh_tw"] = i
        elif h.startswith("17_") or "english" in hl:
            col["en"] = i
        elif h.startswith("30_") or "japanese" in hl or "日本語" in h:
            col["ja"] = i
        elif h == "Image Url" or hl == "image_url":
            col["image_url"] = i
    return col


def _read_translation_sheet(ws, acc: dict) -> None:
    rows = ws.iter_rows(values_only=True)
    try:
        header = [_norm(h) for h in next(rows)]
    except StopIteration:
        return
    col = _resolve_translation_columns(header)
    if "item_id" not in col or "ko" not in col:
        return

    def get(row, key):
        i = col.get(key)
        return row[i] if i is not None and i < len(row) else None

    for row in rows:
        raw_item = get(row, "item_id")
        if raw_item is None or _norm(raw_item) == "":
            continue
        item_id = str(int(raw_item)) if isinstance(raw_item, float) else str(raw_item).strip()
        ko = _norm(get(row, "ko"))
        if ko == "":
            continue
        acc.setdefault(item_id, []).append({
            "org_id": _norm(get(row, "org_id")),
            "ko": ko,
            "en": _norm(get(row, "en")),
            "ja": _norm(get(row, "ja")),
            "zh_cn": _norm(get(row, "zh_cn")),
            "zh_tw": _norm(get(row, "zh_tw")),
        })


def load_fragments(dir_or_file: Path) -> dict:
    """번역본 폴더(여러 .xlsx) 또는 합본 .xlsx -> {item_id(str): [fragment,...]}.

    fragment = {org_id, ko, en, ja, zh_cn, zh_tw}
    """
    acc: dict = {}
    p = Path(dir_or_file)
    files = []
    if p.is_dir():
        # 하위폴더(place별 폴더)까지 재귀 스캔. 엑셀 임시 잠금파일(~$) 제외.
        files = sorted(f for f in p.rglob("*.xlsx") if not f.name.startswith("~$"))
    elif p.is_file():
        files = [p]
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for ws in wb.worksheets:
            _read_translation_sheet(ws, acc)
        wb.close()
    return acc


def _iter_translation_rows(ws):
    """워크시트 -> 행 dict 제너레이터 {item_id, org_id, ko, en, ja, zh_cn, zh_tw, image_url}."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = [_norm(h) for h in next(rows)]
    except StopIteration:
        return
    col = _resolve_translation_columns(header)
    if "ko" not in col:
        return

    def get(row, key):
        i = col.get(key)
        return row[i] if i is not None and i < len(row) else None

    for row in rows:
        ko = _norm(get(row, "ko"))
        if ko == "":
            continue
        raw_item = get(row, "item_id")
        item_id = ""
        if raw_item is not None and _norm(raw_item) != "":
            item_id = str(int(raw_item)) if isinstance(raw_item, float) else str(raw_item).strip()
        yield {
            "item_id": item_id, "org_id": _norm(get(row, "org_id")), "ko": ko,
            "en": _norm(get(row, "en")), "ja": _norm(get(row, "ja")),
            "zh_cn": _norm(get(row, "zh_cn")), "zh_tw": _norm(get(row, "zh_tw")),
            "image_url": _norm(get(row, "image_url")),
        }


def load_uploaded(files: list[tuple[str, bytes]]):
    """업로드된 (상대경로, 바이트) 목록 -> (fragments_by_item, meta_by_item).

    item_id는 파일의 Item Id 컬럼 우선, 없으면 파일명에서. place_id/title은 경로에서.
    image_url은 행의 Image Url 컬럼에서 캡처.
    """
    frags: dict = {}
    meta: dict = {}
    for rel_path, data in files:
        pm = parse_path_meta(rel_path)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            continue
        for ws in wb.worksheets:
            for row in _iter_translation_rows(ws):
                iid = row["item_id"] or pm["item_id"]
                if not iid:
                    continue
                iid = str(iid)
                frags.setdefault(iid, []).append({
                    "org_id": row["org_id"], "ko": row["ko"], "en": row["en"],
                    "ja": row["ja"], "zh_cn": row["zh_cn"], "zh_tw": row["zh_tw"],
                })
                m = meta.setdefault(iid, {"place_id": pm["place_id"], "title": pm["title"],
                                          "image_url": "", "width": None, "height": None})
                if m["place_id"] is None and pm["place_id"] is not None:
                    m["place_id"] = pm["place_id"]
                if not m["title"] and pm["title"]:
                    m["title"] = pm["title"]
                if not m["image_url"] and row["image_url"]:
                    m["image_url"] = row["image_url"]
        wb.close()
    return frags, meta
