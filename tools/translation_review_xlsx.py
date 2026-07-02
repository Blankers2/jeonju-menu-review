"""부가어 번역 검토용 xlsx 생성(자동수정 X). 사람이 확인·반영하도록.

시트:
 - 채우기후보: 번역에 빠진 부가어를 조각으로 채운 제안. **실제로 값이 바뀌는 행만** 출력.
   (조각 탐색: 해당 item 시트 → 같은 place 시트. 전역 최빈 미사용 = 임의번역 금지)
 - 미해결부가어: 조각이 없어 자동으로 못 채우는 부가어(빈도순) → 번역팀/수동.
 - 용량부분누락: 원문 용량이 일부 언어에만 반영된 행.
출력: 번역검토_부가어.xlsx (프로젝트 루트, gitignore)
"""
import glob
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DRAFTS = ROOT / "storage" / "drafts"
sys.path.insert(0, str(ROOT / "tools"))
from fragment_dict import load as load_frag
from translation_fill import modifiers

LANGS = ("en", "ja", "zh_cn", "zh_tw")
_CAP = re.compile(r"(\d+(?:\.\d+)?)\s*(公斤|千克|公升|毫升|克|升|kg|g|ml|l|ℓ|cc)(?![A-Za-z])", re.IGNORECASE)
_UNIT = {"kg": "kg", "g": "g", "ml": "ml", "l": "l", "ℓ": "l", "cc": "ml", "克": "g",
         "公斤": "kg", "千克": "kg", "毫升": "ml", "升": "l", "公升": "l"}


def caps(s):
    return {f"{m.group(1)}{_UNIT.get(m.group(2).lower(), m.group(2).lower())}" for m in _CAP.finditer(s or "")}


def main():
    by_item, _gd = load_frag()

    def frags_of(iids):
        m = {}
        for iid in iids:
            for fr in by_item.get(str(iid), []):
                ko = (fr.get("ko") or "").strip()
                if ko:
                    m.setdefault(ko, {l: (fr.get(l) or "").strip() for l in LANGS})
        return m

    place_items: dict[str, list[str]] = {}
    for f in glob.glob(str(DRAFTS / "*.json")):
        stem = Path(f).stem
        if "_" in stem:
            pid, iid = stem.split("_", 1)
            place_items.setdefault(pid, []).append(iid)

    fill, unresolved, cap_partial = [], Counter(), []
    for f in sorted(glob.glob(str(DRAFTS / "*.json"))):
        d = json.loads(Path(f).read_text(encoding="utf-8"))
        iid = str(d.get("item_id"))
        idict = frags_of([iid])
        pdict = frags_of([x for x in place_items.get(str(d.get("place_id")), []) if x != iid])
        for r in d.get("rows", []):
            menu = r.get("menu") or ""
            for kind, c in modifiers(menu):
                tr = idict.get(c) or pdict.get(c)
                if not tr or not any(tr[l] for l in LANGS):
                    unresolved[c] += 1
                    continue
                row = [d["place_id"], d["item_id"], menu, c]
                changed = False
                for l in LANGS:
                    cur = r.get(l) or ""
                    frag = tr[l]
                    if frag and frag not in cur and cur not in frag:
                        prop = f"{cur} - {frag}" if kind == "suffix" else f"{cur} ({frag})"
                        changed = True
                    else:
                        prop = cur
                    row += [cur, prop.strip()]
                if changed:  # 현재=제안 동일행은 제외(혼란 방지)
                    fill.append(row)
            # 용량 부분누락(일부 언어만)
            cm = caps(menu)
            if cm:
                per = {l: caps(r.get(l) or "") for l in LANGS}
                pres = [l for l in LANGS if (r.get(l) or "") and cm <= per[l]]
                miss = [l for l in LANGS if (r.get(l) or "") and not cm <= per[l]]
                if pres and miss:
                    cap_partial.append([d["place_id"], d["item_id"], menu, ",".join(sorted(cm)),
                                        ",".join(pres), ",".join(miss)])

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "채우기후보"
    ws.append(["※ 번역에 부가어가 빠져 실제로 값이 바뀌는 행만 표시. '제안' 검토 후 반영을 지시하세요."])
    ws.append(["place_id", "item_id", "메뉴명", "부가어",
               "en_현재", "en_제안", "ja_현재", "ja_제안",
               "zh_cn_현재", "zh_cn_제안", "zh_tw_현재", "zh_tw_제안"])
    for row in fill:
        ws.append(row)
    ws2 = wb.create_sheet("미해결부가어")
    ws2.append(["부가어", "건수", "(번역조각 없음 → 번역팀/수동)"])
    for c, n in unresolved.most_common():
        ws2.append([c, n])
    ws3 = wb.create_sheet("용량부분누락")
    ws3.append(["place_id", "item_id", "메뉴명", "용량", "반영언어", "누락언어"])
    for row in cap_partial:
        ws3.append(row)

    out = ROOT / "번역검토_부가어.xlsx"
    wb.save(out)
    print(f"저장: {out}")
    print(f" 채우기후보 {len(fill)}행(실변경만) / 미해결부가어 {len(unresolved)}종 / 용량부분누락 {len(cap_partial)}건")


if __name__ == "__main__":
    main()
